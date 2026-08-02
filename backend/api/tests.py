"""API-layer helper tests that need no database.

Currently targets the pre-match-risk staleness gate: a weekly-load "over
ceiling" verdict must stop firing once its window is no longer recent, so a
stale materialized state can't keep flagging load after a player stopped
training (the "5 días libres → sobreentrenamiento" false positive, §1.4).
"""

from datetime import datetime, timedelta, timezone as _tz

from django.test import SimpleTestCase

from api.command_center import _WEEKLY_LOAD_MAX_STALE_DAYS, _weekly_load_over_ceiling


class WeeklyLoadOverCeilingTests(SimpleTestCase):
    NOW = datetime(2026, 7, 13, 9, 0, tzinfo=_tz.utc)

    def _state(self, status, days_old):
        to = (self.NOW - timedelta(days=days_old)).isoformat()
        return {"weekly_load": {"metrics": [{"status": status}], "window": {"to": to}}}

    def test_fresh_over_is_flagged(self):
        self.assertTrue(_weekly_load_over_ceiling(self._state("over", 0), self.NOW))
        self.assertTrue(_weekly_load_over_ceiling(self._state("over", 2), self.NOW))

    def test_stale_over_is_suppressed(self):
        # The core bug: an "over" verdict on an old window must NOT fire.
        self.assertFalse(_weekly_load_over_ceiling(self._state("over", 5), self.NOW))

    def test_staleness_boundary(self):
        self.assertTrue(
            _weekly_load_over_ceiling(
                self._state("over", _WEEKLY_LOAD_MAX_STALE_DAYS), self.NOW
            )
        )
        self.assertFalse(
            _weekly_load_over_ceiling(
                self._state("over", _WEEKLY_LOAD_MAX_STALE_DAYS + 1), self.NOW
            )
        )

    def test_not_over_is_false(self):
        self.assertFalse(_weekly_load_over_ceiling(self._state("within", 0), self.NOW))

    def test_missing_or_malformed_is_false(self):
        self.assertFalse(_weekly_load_over_ceiling(None, self.NOW))
        self.assertFalse(_weekly_load_over_ceiling({}, self.NOW))
        # "over" but no window
        self.assertFalse(
            _weekly_load_over_ceiling(
                {"weekly_load": {"metrics": [{"status": "over"}]}}, self.NOW
            )
        )
        # "over" but unparseable window.to
        self.assertFalse(
            _weekly_load_over_ceiling(
                {"weekly_load": {"metrics": [{"status": "over"}], "window": {"to": "garbage"}}},
                self.NOW,
            )
        )


class ExportWorkbookTests(SimpleTestCase):
    """The pure workbook builder (§5 export) — no DB."""

    def test_builds_valid_xlsx_with_sheets(self):
        import io as _io

        from openpyxl import load_workbook

        from api.export import workbook_bytes

        data = workbook_bytes([
            {"name": "CK", "headers": ["Jugador", "Fecha", "CK (U/L)"],
             "rows": [["Díaz", "2026-07-10", 720], ["Pérez", "2026-07-10", 310]]},
            {"name": "GPS: partido/sesión?", "headers": ["Jugador"], "rows": [["X"]]},
        ])
        self.assertEqual(data[:2], b"PK")  # xlsx is a zip
        wb = load_workbook(_io.BytesIO(data))
        self.assertEqual(wb.sheetnames[0], "CK")
        self.assertTrue(all(len(n) <= 31 for n in wb.sheetnames))
        self.assertNotIn(":", wb.sheetnames[1])  # invalid char stripped
        ws = wb["CK"]
        self.assertEqual([c.value for c in ws[1]], ["Jugador", "Fecha", "CK (U/L)"])
        self.assertEqual(ws.cell(row=2, column=3).value, 720)

    def test_empty_yields_one_sheet(self):
        import io as _io

        from openpyxl import load_workbook

        from api.export import workbook_bytes

        wb = load_workbook(_io.BytesIO(workbook_bytes([])))
        self.assertEqual(len(wb.sheetnames), 1)


class StatsTests(SimpleTestCase):
    """Pure rolling-statistics helpers (dashboards/stats.py) — §1.2 + §4."""

    def test_mean_stdev_cv(self):
        from dashboards import stats

        self.assertEqual(stats.mean([2, 4, 6]), 4)
        self.assertEqual(stats.stdev([2, 4, 6]), 2.0)
        self.assertEqual(stats.cv([2, 4, 6]), 50.0)
        self.assertIsNone(stats.stdev([5]))  # <2 values
        self.assertIsNone(stats.mean([]))

    def test_ewma(self):
        from dashboards import stats

        self.assertEqual(stats.ewma([1, 1, 1], span=2), 1.0)   # constant series
        self.assertEqual(stats.ewma([0, 10], span=1), 10.0)    # span 1 → latest
        self.assertIsNone(stats.ewma([]))

    def test_deviation_full(self):
        from dashboards import stats

        d = stats.deviation(10, [2, 4, 6])
        self.assertEqual(d["centre"], 4)
        self.assertEqual(d["sd"], 2.0)
        self.assertEqual(d["z"], 3.0)
        self.assertEqual(d["pct"], 150.0)
        self.assertEqual(d["cv"], 50.0)

    def test_deviation_partial_and_empty(self):
        from dashboards import stats

        d = stats.deviation(4, [4, 4])       # zero spread
        self.assertIsNone(d["z"])
        self.assertEqual(d["pct"], 0.0)
        d1 = stats.deviation(5, [5])          # single prior → no sd
        self.assertIsNone(d1["z"])
        self.assertEqual(d1["pct"], 0.0)
        self.assertIsNone(stats.deviation(5, []))          # no prior
        self.assertIsNone(stats.deviation(None, [1, 2, 3]))  # non-numeric value

    def test_deviation_ewma_method(self):
        from dashboards import stats

        d = stats.deviation(10, [2, 4, 6], method="ewma", span=1)
        self.assertEqual(d["centre"], 6.0)   # span-1 EWMA = latest prior
        self.assertEqual(d["z"], 2.0)        # (10-6)/sd(=2)


class ParseDateWindowTests(SimpleTestCase):
    """`_parse_date_window` — the shared date-range sanitizer behind the
    player layout (`/players/{id}/views`) and team reports (`/reports/{slug}`).

    Regression: a date-only upper bound parsed to midnight, so
    `recorded_at__lte=<to>` dropped everything recorded on the final day —
    "últimos 30 días" and custom ranges never showed today's data.
    """

    def _win(self, a, b):
        from api.routers import _parse_date_window

        return _parse_date_window(a, b)

    def test_date_only_upper_bound_covers_the_whole_day(self):
        _, to = self._win("2026-07-03", "2026-08-02")
        self.assertEqual(to, datetime(2026, 8, 2, 23, 59, 59, 999999))

    def test_same_day_range_is_not_empty(self):
        frm, to = self._win("2026-08-02", "2026-08-02")
        self.assertEqual(frm, datetime(2026, 8, 2, 0, 0))
        self.assertEqual(to, datetime(2026, 8, 2, 23, 59, 59, 999999))
        self.assertLess(frm, to)

    def test_lower_bound_stays_at_midnight(self):
        frm, _ = self._win("2026-07-03", "2026-08-02")
        self.assertEqual(frm, datetime(2026, 7, 3, 0, 0))

    def test_explicit_datetime_upper_bound_is_honored(self):
        # An explicit time means the caller meant it — don't stretch it.
        _, to = self._win("2026-07-03", "2026-08-02T09:30:00")
        self.assertEqual(to, datetime(2026, 8, 2, 9, 30))

    def test_inverted_range_swaps_then_stretches_the_upper_bound(self):
        # The stretch must follow the swap, else 23:59:59 lands on the LOWER
        # bound and the range silently loses most of its first day.
        frm, to = self._win("2026-08-05", "2026-08-01")
        self.assertEqual(frm, datetime(2026, 8, 1, 0, 0))
        self.assertEqual(to, datetime(2026, 8, 5, 23, 59, 59, 999999))

    def test_span_cap_still_enforced(self):
        from api.routers import DATE_WINDOW_MAX_DAYS

        frm, to = self._win("2023-08-02", "2026-08-02")
        self.assertEqual((to - frm).days, DATE_WINDOW_MAX_DAYS)

    def test_malformed_and_missing_are_none(self):
        self.assertEqual(self._win("nope", "also-nope"), (None, None))
        self.assertEqual(self._win(None, None), (None, None))
        _, to = self._win(None, "2026-08-02")
        self.assertEqual(to, datetime(2026, 8, 2, 23, 59, 59, 999999))


class UsageBucketStartsTests(SimpleTestCase):
    """`usage_bucket_starts` — the x-axis of the Uso page.

    Regression: the series started at the cutoff and stepped forward, snapping
    each week back to Monday. That shifted the whole run one bucket early, so
    the in-progress week/month was never emitted and results recorded in it
    silently vanished from the chart. The window must always run THROUGH the
    current bucket, present even when there's no data for it.
    """

    def _keys(self, now, bucket, n):
        from api.routers import usage_bucket_starts

        return [d.date().isoformat() for d in usage_bucket_starts(now, bucket, n)]

    # 2026-08-02 is a Sunday; its Monday-anchored week starts 2026-07-27.
    SUNDAY = datetime(2026, 8, 2, 15, 30, tzinfo=_tz.utc)

    def test_weekly_run_ends_on_the_current_week(self):
        keys = self._keys(self.SUNDAY, "week", 12)
        self.assertEqual(len(keys), 12)
        self.assertEqual(keys[-1], "2026-07-27")
        self.assertEqual(keys[0], "2026-05-11")

    def test_sunday_is_not_pushed_into_the_next_week(self):
        # Sunday is weekday()==6 — the boundary the old code got wrong.
        self.assertEqual(self._keys(self.SUNDAY, "week", 1), ["2026-07-27"])

    def test_monday_anchors_on_itself(self):
        monday = datetime(2026, 7, 27, 0, 5, tzinfo=_tz.utc)
        self.assertEqual(self._keys(monday, "week", 1), ["2026-07-27"])

    def test_weekly_buckets_are_consecutive_mondays(self):
        keys = self._keys(self.SUNDAY, "week", 6)
        days = [datetime.fromisoformat(k) for k in keys]
        self.assertTrue(all(d.weekday() == 0 for d in days))
        gaps = {(b - a).days for a, b in zip(days, days[1:])}
        self.assertEqual(gaps, {7})

    def test_monthly_run_ends_on_the_current_month(self):
        keys = self._keys(self.SUNDAY, "month", 6)
        self.assertEqual(keys, [
            "2026-03-01", "2026-04-01", "2026-05-01",
            "2026-06-01", "2026-07-01", "2026-08-01",
        ])

    def test_monthly_does_not_skip_february(self):
        # The 31-day step turned Jan 31 into Mar 3, dropping February.
        keys = self._keys(datetime(2026, 3, 31, 12, 0, tzinfo=_tz.utc), "month", 3)
        self.assertEqual(keys, ["2026-01-01", "2026-02-01", "2026-03-01"])

    def test_monthly_crosses_the_year_boundary(self):
        keys = self._keys(datetime(2026, 2, 10, 12, 0, tzinfo=_tz.utc), "month", 4)
        self.assertEqual(keys, ["2025-11-01", "2025-12-01", "2026-01-01", "2026-02-01"])

    def test_bucket_starts_are_midnight(self):
        from api.routers import usage_bucket_starts

        for b in ("week", "month"):
            for d in usage_bucket_starts(self.SUNDAY, b, 3):
                self.assertEqual((d.hour, d.minute, d.second, d.microsecond), (0, 0, 0, 0))


from django.contrib.auth import get_user_model  # noqa: E402
from django.test import RequestFactory, TestCase  # noqa: E402
from django.utils import timezone  # noqa: E402


class TeamWidgetArrangeTests(TestCase):
    """§2.c panel-builder arrange endpoints (called directly with a superuser
    request — bypasses ninja auth but exercises require_perm + the logic)."""

    def setUp(self):
        from core.models import Category, Club, Department
        from dashboards.models import (
            ChartType, TeamReportLayout, TeamReportSection, TeamReportWidget,
        )
        self.rf = RequestFactory()
        self.su = get_user_model().objects.create_superuser("su", "su@x.com", "x")
        self.club = Club.objects.create(name="FC")
        self.dept = Department.objects.create(club=self.club, name="Med", slug="med")
        self.cat = Category.objects.create(club=self.club, name="A")
        self.cat.departments.add(self.dept)
        self.layout = TeamReportLayout.objects.create(department=self.dept, category=self.cat)
        self.section = TeamReportSection.objects.create(layout=self.layout)
        self.w1 = TeamReportWidget.objects.create(
            section=self.section, chart_type=ChartType.TEAM_LEADERBOARD.value,
            title="A", sort_order=0, column_span=6,
        )
        self.w2 = TeamReportWidget.objects.create(
            section=self.section, chart_type=ChartType.TEAM_LEADERBOARD.value,
            title="B", sort_order=1, column_span=6,
        )

    def _req(self, method):
        r = getattr(self.rf, method)("/x")
        r.user = self.su
        return r

    def test_resize_clamps_span(self):
        from api.routers import WidgetArrangeIn, update_team_widget
        update_team_widget(self._req("patch"), str(self.w1.id), WidgetArrangeIn(column_span=99))
        self.w1.refresh_from_db()
        self.assertEqual(self.w1.column_span, 12)

    def test_reorder_reassigns_sort_order(self):
        from api.routers import WidgetReorderIn, reorder_team_widgets
        reorder_team_widgets(
            self._req("post"),
            WidgetReorderIn(widget_ids=[str(self.w2.id), str(self.w1.id)]),
        )
        self.w1.refresh_from_db()
        self.w2.refresh_from_db()
        self.assertEqual(self.w2.sort_order, 0)
        self.assertEqual(self.w1.sort_order, 1)

    def test_delete_removes_widget(self):
        from api.routers import delete_team_widget
        from dashboards.models import TeamReportWidget
        delete_team_widget(self._req("delete"), str(self.w1.id))
        self.assertFalse(TeamReportWidget.objects.filter(id=self.w1.id).exists())


class EpisodeAvailableAtTests(TestCase):
    """§3.1 — Episode.available_at write path via PATCH /episodes/{id}."""

    def setUp(self):
        from django.utils import timezone
        from core.models import Category, Club, Department, Player
        from exams.models import Episode, ExamTemplate
        self.rf = RequestFactory()
        self.su = get_user_model().objects.create_superuser("su2", "su2@x.com", "x")
        self.club = Club.objects.create(name="FC2")
        self.dept = Department.objects.create(club=self.club, name="Med", slug="med")
        self.cat = Category.objects.create(club=self.club, name="A")
        self.player = Player.objects.create(category=self.cat, first_name="A", last_name="B")
        self.template = ExamTemplate.objects.create(
            name="Lesiones", slug="lesiones", department=self.dept, is_episodic=True,
            episode_config={"stage_field": "stage", "open_stages": ["injured"], "closed_stage": "closed"},
            config_schema={"fields": [{"key": "stage", "type": "categorical", "options": ["injured", "closed"]}]},
        )
        self.ep = Episode.objects.create(
            player=self.player, template=self.template, status="open",
            stage="injured", started_at=timezone.now(),
        )

    def _req(self):
        r = self.rf.patch("/x")
        r.user = self.su
        return r

    def test_set_and_serialize_available_at(self):
        from api.routers import update_episode
        from api.schemas import EpisodePatchIn
        out = update_episode(self._req(), str(self.ep.id), EpisodePatchIn(available_at="2026-06-01"))
        self.ep.refresh_from_db()
        self.assertIsNotNone(self.ep.available_at)
        self.assertEqual(out["available_at"].date().isoformat(), "2026-06-01")

    def test_clear_available_at(self):
        from api.routers import update_episode
        from api.schemas import EpisodePatchIn
        update_episode(self._req(), str(self.ep.id), EpisodePatchIn(available_at="2026-06-01"))
        update_episode(self._req(), str(self.ep.id), EpisodePatchIn(available_at="clear"))
        self.ep.refresh_from_db()
        self.assertIsNone(self.ep.available_at)

    def test_setting_available_at_does_not_close_episode(self):
        from api.routers import update_episode
        from api.schemas import EpisodePatchIn
        update_episode(self._req(), str(self.ep.id), EpisodePatchIn(available_at="2026-06-01"))
        self.ep.refresh_from_db()
        self.assertEqual(self.ep.status, "open")  # available ≠ closed
        self.assertIsNone(self.ep.ended_at)


class ForecastAccuracyTests(TestCase):
    """§3.2 — return-prognosis accuracy: error = actual − first forecast."""

    def _episode(self, first_expected, actual_iso):
        from datetime import datetime, timezone as _tzc
        from exams.models import Episode, ExamResult
        ep = Episode.objects.create(
            player=self.player, template=self.template, status="closed",
            stage="closed", started_at=timezone.now(),
            available_at=datetime.fromisoformat(f"{actual_iso}T12:00:00+00:00"),
        )
        ExamResult.objects.create(
            player=self.player, template=self.template, episode=ep,
            recorded_at=timezone.now(),
            result_data={"expected_return_date": first_expected},
        )
        return ep

    def setUp(self):
        from django.utils import timezone as _tz
        from core.models import Category, Club, Department, Player
        from exams.models import ExamTemplate
        self.club = Club.objects.create(name="FC3")
        self.dept = Department.objects.create(club=self.club, name="Med", slug="med")
        self.cat = Category.objects.create(club=self.club, name="A")
        self.player = Player.objects.create(category=self.cat, first_name="A", last_name="B")
        self.template = ExamTemplate.objects.create(
            name="Lesiones", slug="lesiones", department=self.dept, is_episodic=True,
            episode_config={"stage_field": "stage", "open_stages": ["injured"], "closed_stage": "closed"},
            config_schema={"fields": [{"key": "stage", "type": "categorical", "options": ["injured", "closed"]}]},
        )
        self.template.applicable_categories.add(self.cat)

    def test_bias_and_mae(self):
        from api.injury_forecast import forecast_accuracy
        self._episode("2026-06-30", "2026-07-10")   # +10 (late)
        self._episode("2026-06-20", "2026-06-15")   # −5 (early)
        out = forecast_accuracy(category=self.cat, department=self.dept)
        self.assertEqual(out["episodes"], 2)
        self.assertEqual(out["bias_days"], 2.5)      # (10 + −5) / 2
        self.assertEqual(out["mae_days"], 7.5)       # (10 + 5) / 2
        # worst-first ordering
        self.assertEqual(out["samples"][0]["error_days"], 10)

    def test_episode_without_forecast_is_skipped(self):
        from exams.models import Episode
        from api.injury_forecast import forecast_accuracy
        # available but no expected_return_date ever recorded → not scored.
        Episode.objects.create(
            player=self.player, template=self.template, status="closed",
            stage="closed", started_at=timezone.now(),
            available_at=timezone.now(),
        )
        self.assertEqual(forecast_accuracy(category=self.cat)["episodes"], 0)
