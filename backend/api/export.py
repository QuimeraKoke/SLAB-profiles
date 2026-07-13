"""Raw-data export to Excel (§5 — "el dato pertenece al club").

One workbook, one **sheet per exam type**, each row = (jugador, fecha) with all
that exam's field values — calculated fields included (they already live in
`result_data`). Server-side via openpyxl. Self-service: filters by category,
players, exams (by department), date range and session type. Access is gated by
JWT + the standard scoping (club / category / department); the menu action is
shown to Editors frontend-side.
"""

from __future__ import annotations

import io
import re
from typing import Any

# Excel sheet-title rules: ≤31 chars, none of : \ / ? * [ ]
_INVALID_SHEET = re.compile(r"[:\\/?*\[\]]")


def _sheet_title(name: str, used: set[str]) -> str:
    base = (_INVALID_SHEET.sub(" ", name or "Hoja").strip() or "Hoja")[:31]
    title, i = base, 2
    while title.lower() in used:
        suffix = f" ({i})"
        title = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(title.lower())
    return title


def workbook_bytes(sheets: list[dict]) -> bytes:
    """Build an .xlsx from `[{"name", "headers": [str], "rows": [[cell,…]]}]`.

    Pure (no DB) so it's unit-testable. Always yields ≥1 sheet.
    """
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    used: set[str] = set()
    if not sheets:
        wb.create_sheet(_sheet_title("Sin datos", used))
    for s in sheets:
        ws = wb.create_sheet(_sheet_title(s.get("name", "Hoja"), used))
        ws.append(list(s.get("headers") or []))
        for row in s.get("rows") or []:
            ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header(field: dict) -> str:
    label = field.get("label") or field.get("key") or ""
    unit = field.get("unit")
    return f"{label} ({unit})" if unit else label


def _humanize(value: Any, field: dict) -> Any:
    """Categorical → its display label; everything else as-is; None → ""."""
    if value is None:
        return ""
    if field.get("type") == "categorical" and isinstance(value, str):
        return (field.get("option_labels") or {}).get(value, value)
    return value


def build_export(*, category, membership, template_ids=None, player_ids=None,
                 date_from=None, date_to=None, event_type=None) -> bytes:
    """Query the scoped results and render the workbook (one sheet per exam)."""
    from django.utils import timezone

    from core.models import Player
    from exams.models import ExamResult, ExamTemplate
    from api.scoping import scope_players, scope_results, scope_templates
    from dashboards.assistant_tools import _schema_fields

    templates = scope_templates(
        ExamTemplate.objects.select_related("department")
        .filter(applicable_categories=category),
        membership,
    )
    if template_ids:
        templates = templates.filter(id__in=template_ids)
    templates = list(templates.order_by("department__name", "name").distinct())

    players = scope_players(
        Player.objects.filter(category=category).select_related("position"),
        membership,
    )
    if player_ids:
        players = players.filter(id__in=player_ids)
    pmeta = {
        p.id: (
            f"{p.first_name} {p.last_name}".strip(),
            p.position.abbreviation if p.position else "",
        )
        for p in players
    }
    pids = list(pmeta.keys())

    sheets: list[dict] = []
    for t in templates:
        fields = _schema_fields(t)
        headers = ["Jugador", "Posición", "Fecha", "Tipo sesión"] + [
            _header(f) for f in fields
        ]
        qs = scope_results(
            ExamResult.objects.filter(template=t, player_id__in=pids)
            .select_related("event"),
            membership,
        )
        if date_from:
            qs = qs.filter(recorded_at__gte=date_from)
        if date_to:
            qs = qs.filter(recorded_at__lte=date_to)
        if event_type:
            qs = qs.filter(event__event_type=event_type)

        rows = []
        for r in qs.order_by("player_id", "recorded_at"):
            name, pos = pmeta.get(r.player_id, ("", ""))
            data = r.result_data or {}
            sess = (
                (r.event.event_type if r.event_id else None)
                or data.get("tipo_sesion")
                or ""
            )
            row = [
                name, pos,
                timezone.localtime(r.recorded_at).date().isoformat(), sess,
            ]
            row.extend(_humanize(data.get(f.get("key")), f) for f in fields)
            rows.append(row)

        sheets.append({"name": t.name, "headers": headers, "rows": rows})

    return workbook_bytes(sheets)
