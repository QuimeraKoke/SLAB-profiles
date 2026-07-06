"""Import the club's "Histórico Fatiga Central" workbook into `fatiga_central`.

Reads the long-format 'Datos' sheet (one row per player per session: Fecha,
Jugador, I1/I2/I3 in Hz, CFF basal, PR, EA). The sheet's computed columns
(CFF mean, Δ%, Var %) are NOT trusted — occasional cells carry spreadsheet
artifacts (dates where numbers belong) — so the importer recomputes them from
the raw I1–I3 readings with the same formulas the template uses. Per-row
basals that are non-numeric fall back to the player's most common basal in
the file.

Additive + idempotent: one ExamResult per (player, session date); rows whose
(player, date) already exist are skipped, so re-running with an updated
workbook only adds new sessions. Player matching is alias-first, then
first-name token + last-name token (accent-insensitive), which absorbs the
sheet's maternal-surname / full-caps labels ("FABIAN HORMAZABAL BERRIOS" →
Fabián Hormazábal). Unmatched names are reported and skipped, never invented.

The workbook ships as .ods — convert to .xlsx first (the club's own workflow
already exports .xlsx for Google Sheets):

    docker compose exec backend python manage.py import_fatiga_central \\
        --file /tmp/fatiga_central.xlsx            # dry-run: show the plan
    docker compose exec backend python manage.py import_fatiga_central \\
        --file /tmp/fatiga_central.xlsx --commit   # write
"""
from __future__ import annotations

import json
import unicodedata
from collections import Counter
from datetime import date, datetime, time
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import Category, Player, PlayerAlias
from exams.models import ExamResult, ExamTemplate


def _norm(s: str | None) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower().strip()


def _to_date(raw) -> date | None:
    if raw in (None, ""):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    try:
        return datetime.fromisoformat(str(raw).strip()[:19]).date()
    except ValueError:
        return None


def _to_float(raw) -> float | None:
    """Parse a numeric cell; date/datetime artifacts and '—' read as None."""
    if raw in (None, ""):
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (datetime, date)):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    try:
        return float(str(raw).replace(",", ".").strip())
    except ValueError:
        return None


class Command(BaseCommand):
    help = "Import the Histórico Fatiga Central workbook into the 'fatiga_central' template."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the .xlsx workbook.")
        parser.add_argument("--club", default="Universidad de Chile")
        parser.add_argument("--category", default="Primer Equipo")
        parser.add_argument("--sheet", default="Datos")
        parser.add_argument("--commit", action="store_true",
                            help="Write to the database (default is a dry-run plan).")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        category = Category.objects.filter(
            club__name=opts["club"], name=opts["category"],
        ).select_related("club").first()
        if category is None:
            raise CommandError(f"Category '{opts['category']}' not found in club '{opts['club']}'.")
        template = ExamTemplate.objects.filter(
            slug="fatiga_central", department__club=category.club, is_active_version=True,
        ).first()
        if template is None:
            raise CommandError("Template 'fatiga_central' not found — run seed_fatiga_central first.")

        wb = load_workbook(path, data_only=True)
        if opts["sheet"] not in wb.sheetnames:
            raise CommandError(f"Sheet '{opts['sheet']}' not in workbook ({wb.sheetnames}).")
        ws = wb[opts["sheet"]]

        # The header row floats below a title block — find it by content.
        header_row, col = None, {}
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if "Fecha" in cells and "Jugador" in cells:
                header_row = row_idx
                col = {h: i for i, h in enumerate(cells) if h}
                break
        if header_row is None:
            raise CommandError("Could not find the header row (needs 'Fecha' and 'Jugador' columns).")
        for required in ("Fecha", "Jugador", "I1 (Hz)", "I2 (Hz)", "I3 (Hz)"):
            if required not in col:
                raise CommandError(f"Missing column '{required}' (found: {sorted(col)})")

        # Includes INACTIVE players on purpose — historical sessions belong
        # on preserved records of ex-players too.
        players = list(Player.objects.filter(category=category))
        aliases = {
            _norm(a.value): a.player_id
            for a in PlayerAlias.objects.filter(player__category=category)
        }
        by_id = {p.id: p for p in players}

        def match_player(raw_name: str) -> Player | None:
            pid = aliases.get(_norm(raw_name))
            if pid and pid in by_id:
                return by_id[pid]
            toks = _norm(raw_name).split()
            if len(toks) < 2:
                return None
            cands = []
            for p in players:
                first_toks = _norm(p.first_name).split()
                last_join = _norm(p.last_name).replace(" ", "")
                if not first_toks or toks[0] != first_toks[0]:
                    continue
                if any(t == last_join for t in toks[1:]):
                    cands.append(p)
            return cands[0] if len(cands) == 1 else None

        # Existing (player, day) keys — recorded_at date is the session key.
        existing: set[tuple] = set()
        for pid, rec in ExamResult.objects.filter(
            template__family_id=template.family_id, player__category=category,
        ).values_list("player_id", "recorded_at"):
            existing.add((pid, timezone.localtime(rec).date().isoformat()))

        # First pass — parse rows, collect per-player basal candidates.
        rows, unmatched, bad = [], Counter(), 0
        basal_seen: dict = {}
        for idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
            if not any(c not in (None, "") for c in row):
                continue
            name = str(row[col["Jugador"]] or "").strip()
            day = _to_date(row[col["Fecha"]])
            i_vals = [_to_float(row[col[k]]) for k in ("I1 (Hz)", "I2 (Hz)", "I3 (Hz)")]
            present = [v for v in i_vals if v is not None]
            if not name or day is None or not present:
                bad += 1
                continue
            player = match_player(name)
            if player is None:
                unmatched[name] += 1
                continue
            basal = _to_float(row[col["CFF basal"]]) if "CFF basal" in col else None
            if basal is not None:
                basal_seen.setdefault(player.id, Counter())[basal] += 1
            rows.append({
                "row": idx, "player": player, "day": day, "i": i_vals,
                "basal": basal,
                "pr": _to_float(row[col["PR"]]) if "PR" in col else None,
                "ea": _to_float(row[col["EA"]]) if "EA" in col else None,
            })

        # Second pass — recompute derived metrics, fill basal gaps, dedupe.
        plan, skipped_dup = [], 0
        for r in sorted(rows, key=lambda r: (r["day"], r["player"].last_name)):
            key = (r["player"].id, r["day"].isoformat())
            if key in existing:
                skipped_dup += 1
                continue
            existing.add(key)

            present = [v for v in r["i"] if v is not None]
            mean = round(sum(present) / len(present), 2)
            basal = r["basal"]
            if basal is None and r["player"].id in basal_seen:
                basal = basal_seen[r["player"].id].most_common(1)[0][0]

            data: dict = {
                k: v for k, v in zip(("i1", "i2", "i3"), r["i"]) if v is not None
            }
            data["cff_mean"] = mean
            if basal is not None:
                data["cff_basal"] = basal
                data["delta_basal_pct"] = round((mean / basal - 1) * 100, 2)
            if len(present) >= 2:
                data["var_intra_pct"] = round((max(present) - min(present)) / mean * 100, 2)
            for k in ("pr", "ea"):
                v = r[k]
                if v is not None:
                    data[k] = int(v) if float(v).is_integer() else v
            plan.append({"row": r["row"], "player": r["player"], "day": r["day"], "data": data})

        self.stdout.write(
            f"Plan: crear {len(plan)} resultados CFF · ya existentes {skipped_dup} · "
            f"filas inválidas {bad} · nombres sin match {sum(unmatched.values())}"
        )
        for name, n in unmatched.most_common():
            self.stdout.write(self.style.WARNING(f"  sin match: {name} ({n} filas) — omitido"))
        per_player: Counter = Counter(p["player"].last_name for p in plan)
        self.stdout.write("  por jugador: " + ", ".join(
            f"{k} {v}" for k, v in sorted(per_player.items())))

        if not opts["commit"]:
            self.stdout.write(self.style.NOTICE("Dry-run — nada escrito. Repite con --commit."))
            return

        with transaction.atomic():
            ExamResult.objects.bulk_create([
                ExamResult(
                    player=p["player"], template=template,
                    recorded_at=timezone.make_aware(datetime.combine(p["day"], time(12, 0))),
                    result_data=p["data"],
                    inputs_snapshot={},
                )
                for p in plan
            ], batch_size=400)

        # Band alerts off each affected player's LATEST reading only —
        # historical rows must not fire (nor mask) current-state alerts.
        from goals.evaluator import evaluate_threshold_rules_for_result
        fired = 0
        for pid in {p["player"].id for p in plan}:
            latest = ExamResult.objects.filter(
                player_id=pid, template=template,
            ).order_by("-recorded_at").first()
            if latest is not None:
                fired += len(evaluate_threshold_rules_for_result(latest))

        log_dir = Path(__file__).resolve().parents[3] / "migration_runs"
        log_dir.mkdir(exist_ok=True)
        log_path = log_dir / f"fatiga-central-import-{timezone.now().strftime('%Y%m%dT%H%M%S')}.jsonl"
        with log_path.open("w") as fh:
            fh.write(json.dumps({"kind": "header", "file": path.name, "created": len(plan),
                                 "skipped_existing": skipped_dup, "unmatched": dict(unmatched)}) + "\n")
            for p in plan:
                fh.write(json.dumps({
                    "row": p["row"], "player": f"{p['player'].first_name} {p['player'].last_name}",
                    "fecha": p["day"].isoformat(), **p["data"],
                }, ensure_ascii=False) + "\n")

        self.stdout.write(self.style.SUCCESS(
            f"Listo: {len(plan)} resultados CFF creados · {fired} alertas de banda "
            f"evaluadas sobre las últimas tomas. Log: {log_path.name}"
        ))
