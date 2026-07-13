"""Import a ForceFrame "Hip AD/AB" export (xlsx) into the hip_adab template.

Export shape (VALD ForceFrame): one row per (player, session, Direction)
where Direction ∈ {Pull, Squeeze}; the pair collapses into ONE ExamResult:

    Name | ExternalId | Date | Time | Device | Mode | Test | Direction |
    Position | L Reps | R Reps | L Max Force (N) | R Max Force (N)

Not wired through `bulk_ingest` because that pipeline takes ONE recorded_at
for the whole file, while these exports span several session dates.

Players are matched by normalized name tokens against the club's ACTIVE
players (all categories — youth train with the first team). Ambiguous or
unknown names abort the run before anything is written.

Idempotent: a (player, date) pair that already has a hip_adab result for
that calendar day is skipped, so re-running the same export is safe.

Run (dry-run first; add --commit to write):

    docker compose exec backend python manage.py import_hip_adab \\
        --file /tmp/hip_adab.xlsx --club "Universidad de Chile" --commit
"""

from __future__ import annotations

import unicodedata
from datetime import datetime, time as time_cls

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import Player
from exams.calculations import compute_result_data
from exams.models import ExamResult, ExamTemplate


def _norm(s: str) -> list[str]:
    s = unicodedata.normalize("NFD", (s or "").upper())
    return "".join(c for c in s if not unicodedata.combining(c)).split()


class Command(BaseCommand):
    help = "Import a ForceFrame Hip AD/AB xlsx export into the hip_adab template."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--club", default="Universidad de Chile")
        parser.add_argument("--commit", action="store_true",
                            help="Write results. Without it: dry-run report only.")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        template = ExamTemplate.objects.filter(
            slug="hip_adab", department__club__name__icontains=opts["club"],
        ).first()
        if template is None:
            raise CommandError(f"Template 'hip_adab' not found for club '{opts['club']}'.")

        try:
            ws = load_workbook(opts["file"], data_only=True).worksheets[0]
        except Exception as exc:  # noqa: BLE001 — surface any open/parse problem
            raise CommandError(f"Cannot read '{opts['file']}': {exc}")

        headers = [str(c).strip() if c is not None else "" for c in
                   next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col = {h: i for i, h in enumerate(headers)}
        required = ["Name", "Date", "Direction", "L Max Force (N)", "R Max Force (N)"]
        missing = [h for h in required if h not in col]
        if missing:
            raise CommandError(f"Missing column(s): {', '.join(missing)} — got {headers}")

        # ---- group rows by (name, date) ---------------------------------
        groups: dict[tuple[str, object], dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[col["Name"]] or "").strip()
            if not name:
                continue
            day = row[col["Date"]]
            day = day.date() if isinstance(day, datetime) else day
            g = groups.setdefault((name, day), {"time": row[col.get("Time", -1)] if "Time" in col else None})
            direction = str(row[col["Direction"]] or "").strip().lower()
            if direction not in ("pull", "squeeze"):
                raise CommandError(f"Unknown Direction {row[col['Direction']]!r} for {name} {day}.")
            g[f"{direction}_left_max"] = row[col["L Max Force (N)"]]
            g[f"{direction}_right_max"] = row[col["R Max Force (N)"]]

        # ---- match players (club-wide actives — youth train with the 1st team)
        roster = list(Player.objects.filter(
            category__club__name__icontains=opts["club"], is_active=True,
        ).select_related("category"))
        resolved: dict[str, Player] = {}
        problems: list[str] = []
        for name in sorted({n for n, _ in groups}):
            tokens = _norm(name)
            cands = [p for p in roster
                     if all(t in tokens for t in _norm(f"{p.first_name} {p.last_name}"))]
            if len(cands) == 1:
                resolved[name] = cands[0]
            elif not cands:
                problems.append(f"sin match: {name!r}")
            else:
                names = ", ".join(f"{p.first_name} {p.last_name} ({p.category.name})" for p in cands)
                problems.append(f"ambiguo: {name!r} -> {names}")
        if problems:
            raise CommandError("Jugadores no resueltos:\n  " + "\n  ".join(problems))

        # ---- build + (optionally) write ---------------------------------
        tz = timezone.get_default_timezone()
        created = skipped = incomplete = 0
        with transaction.atomic():
            for (name, day), g in sorted(groups.items(), key=lambda kv: kv[0][1]):
                player = resolved[name]
                raw = {k: g.get(k) for k in
                       ("pull_left_max", "pull_right_max", "squeeze_left_max", "squeeze_right_max")}
                if any(v is None for v in raw.values()):
                    incomplete += 1
                    self.stdout.write(self.style.WARNING(
                        f"  incompleto (falta Pull o Squeeze): {name} {day} — saltado"))
                    continue
                if ExamResult.objects.filter(
                    template__family_id=template.family_id, player=player,
                    recorded_at__date=day,
                ).exists():
                    skipped += 1
                    continue
                t = g.get("time") if isinstance(g.get("time"), time_cls) else time_cls(12, 0)
                recorded_at = timezone.make_aware(datetime.combine(day, t), tz)
                data, snapshot = compute_result_data(template, raw, player=player)
                line = (f"  {player.first_name} {player.last_name:<22} {day} | "
                        f"pull {raw['pull_left_max']}/{raw['pull_right_max']} "
                        f"sq {raw['squeeze_left_max']}/{raw['squeeze_right_max']} | "
                        f"imb P {data.get('pull_imbalance'):+.1f}% S {data.get('squeeze_imbalance'):+.1f}%")
                self.stdout.write(line)
                if opts["commit"]:
                    # .create() (not bulk_create) so post_save signals run:
                    # player-state writeback + alert-rule evaluation.
                    ExamResult.objects.create(
                        player=player, template=template, recorded_at=recorded_at,
                        result_data=data, inputs_snapshot=snapshot,
                    )
                created += 1

        mode = "CREATED" if opts["commit"] else "would create (dry-run)"
        self.stdout.write(self.style.SUCCESS(
            f"{mode}: {created} | skipped (already imported): {skipped} | incomplete: {incomplete}"
        ))
