"""Import the Pentacompartimental (5-component anthropometry) xlsx export into
the `pentacompartimental` template.

Reads the "Modelo 5 componentes" sheet (blocks per player: a name in col 1
followed by one row per measurement date). Maps the raw ISAK measurements to
the template's raw field keys, matches names to U. de Chile players (accent-
and prefix-aware token matching, across ALL categories so youth players listed
with the first team still resolve), dedups on (player, measurement date) so
existing results are NEVER overwritten, recomputes the 5 masses + indices via
the template formulas, and bulk-creates the new results (signal-free, so no
alert/state recompute flood). Dry-run by default.

    # copy the workbook into the backend mount so the container can read it:
    cp report.xlsx backend/_penta_import.xlsx
    docker compose exec backend python manage.py import_pentacompartimental \\
        --file /app/_penta_import.xlsx            # dry-run
    docker compose exec backend python manage.py import_pentacompartimental \\
        --file /app/_penta_import.xlsx --commit   # write
"""

from __future__ import annotations

import unicodedata
from datetime import datetime

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import Club, Player
from exams.calculations import compute_result_data
from exams.models import ExamResult, ExamTemplate

SHEET = "Modelo 5 componentes"

# Excel column (1-based) → template raw field key.
RAW_MAP = {
    4: "peso", 5: "talla", 6: "talla_sentado",
    7: "biacromial", 8: "diam_torax_transverso", 9: "diam_torax_ap",
    10: "bi_iliocrestideo", 11: "humero", 12: "femur",
    13: "perim_cabeza", 14: "perim_brazo_relajado", 15: "perim_brazo_contraido",
    16: "perim_antebrazo", 17: "perim_torax", 18: "cintura", 19: "caderas",
    20: "muslo_gluteo", 21: "muslo_medio", 22: "pierna_perim",
    23: "pliegue_bicipital", 24: "pliegue_triceps", 25: "pliegue_subescapular",
    26: "pliegue_supracrestideo", 27: "pliegue_supra", 28: "pliegue_abdomen",
    29: "pliegue_muslo", 30: "pliegue_pierna",
}


def _tokens(s: str) -> list[str]:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return [t for t in s.replace(".", " ").replace(",", " ").split() if len(t) > 1]


def _tok_match(a: str, b: str) -> bool:
    """Two name tokens match if equal or one is a ≥4-char prefix of the other
    (handles Bolaño/Bolaños, etc.)."""
    if a == b:
        return True
    if len(a) >= 4 and b.startswith(a):
        return True
    if len(b) >= 4 and a.startswith(b):
        return True
    return False


def _num(v):
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def _to_date(v):
    if v in (None, ""):
        return None
    if hasattr(v, "date"):
        return v.date()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


class Command(BaseCommand):
    help = "Import the Pentacompartimental anthropometry xlsx (additive, dedup on player+date)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the .xlsx (inside the container).")
        parser.add_argument("--club", default="Universidad de Chile")
        parser.add_argument("--sheet", default=SHEET)
        parser.add_argument("--slug", default="pentacompartimental")
        parser.add_argument("--commit", action="store_true", help="Write (default: dry-run).")

    def handle(self, *args, **opts):
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError:
            raise CommandError("openpyxl is required (pip install openpyxl).")

        club = Club.objects.filter(name=opts["club"]).first()
        if club is None:
            raise CommandError(f"Club '{opts['club']}' not found.")
        tpl = ExamTemplate.objects.filter(
            slug=opts["slug"], department__club=club,
        ).first()
        if tpl is None:
            raise CommandError(f"Template '{opts['slug']}' not found in {club.name}.")

        # Candidate players: ALL of the club (any category), so youth players
        # listed with the first team still resolve. Prefer Primer Equipo on ties.
        cand = [
            (p, _tokens(f"{p.first_name} {p.last_name}"))
            for p in Player.objects.filter(category__club=club).select_related("category")
        ]

        def match(name: str):
            et = _tokens(name)
            best, best_score = None, (-1, -1)
            for p, pt in cand:
                score = sum(1 for a in et if any(_tok_match(a, b) for b in pt))
                pe = 1 if p.category and p.category.name == "Primer Equipo" else 0
                if (score, pe) > best_score:
                    best_score, best = (score, pe), p
            return best if best_score[0] >= 2 else None

        wb = openpyxl.load_workbook(opts["file"], data_only=True, read_only=True)
        if opts["sheet"] not in wb.sheetnames:
            raise CommandError(f"Sheet '{opts['sheet']}' not in {wb.sheetnames}.")
        ws = wb[opts["sheet"]]
        rows = list(ws.iter_rows(min_row=6, values_only=True))

        # Existing (player_id, date) so we never overwrite.
        existing: set[tuple] = set()
        matched_ids = set()  # fill after matching to scope the query — but we
        # don't know them yet; query broadly for this template instead.
        for r in ExamResult.objects.filter(template=tpl).values_list("player_id", "recorded_at"):
            existing.add((r[0], r[1].date()))

        per_player: dict = {}
        unmatched: dict[str, int] = {}
        to_create: list[ExamResult] = []
        seen_in_file: set[tuple] = set()
        cur_name = None
        skipped_existing = 0
        skipped_dupe = 0

        for r in rows:
            if r[0] not in (None, ""):
                cur_name = str(r[0]).strip()
            if cur_name is None:
                continue
            d = _to_date(r[1])
            if d is None:
                continue
            player = match(cur_name)
            if player is None:
                unmatched[cur_name] = unmatched.get(cur_name, 0) + 1
                continue
            key = (player.id, d)
            if key in existing:
                skipped_existing += 1
                continue
            if key in seen_in_file:
                skipped_dupe += 1
                continue
            seen_in_file.add(key)
            matched_ids.add(player.id)

            raw = {k: _num(r[c - 1]) for c, k in RAW_MAP.items() if _num(r[c - 1]) is not None}
            if not raw.get("peso") or not raw.get("talla"):
                # Guard against empty/garbage rows.
                continue
            raw["sexo"] = 1 if (getattr(player, "sex", "M") or "M").upper().startswith("M") else 2
            result_data, inputs_snapshot = compute_result_data(tpl, raw, player=player)
            recorded_at = timezone.make_aware(
                datetime.combine(d, datetime.min.time()), timezone.get_default_timezone()
            )
            to_create.append(ExamResult(
                player=player, template=tpl, recorded_at=recorded_at,
                result_data=result_data, inputs_snapshot=inputs_snapshot,
            ))
            slot = per_player.setdefault(
                player.id, {"name": cur_name, "player": f"{player.first_name} {player.last_name}",
                            "cat": player.category.name if player.category else "?", "new": 0})
            slot["new"] += 1

        # ---- report ----
        self.stdout.write(f"Template: {tpl.name} ({club.name}) | filas leídas: {len(rows)}")
        self.stdout.write(
            f"Nuevos a crear: {len(to_create)} | ya existían (omitidos): {skipped_existing} "
            f"| dup en archivo: {skipped_dupe} | jugadores: {len(per_player)}"
        )
        for pid, s in sorted(per_player.items(), key=lambda kv: kv[1]["player"]):
            self.stdout.write(f"  +{s['new']:>2}  {s['player']} [{s['cat']}]  (excel: {s['name']})")
        if unmatched:
            self.stdout.write(self.style.WARNING(
                f"\nSIN MATCH ({len(unmatched)}): " + ", ".join(f"{n} ({c})" for n, c in unmatched.items())
            ))

        if not opts["commit"]:
            self.stdout.write(self.style.WARNING("\nDRY-RUN — nada escrito. Pasá --commit para importar."))
            return

        with transaction.atomic():
            ExamResult.objects.bulk_create(to_create, batch_size=500)
        self.stdout.write(self.style.SUCCESS(
            f"\nImportado (aditivo, sin señales): +{len(to_create)} resultados."
        ))
