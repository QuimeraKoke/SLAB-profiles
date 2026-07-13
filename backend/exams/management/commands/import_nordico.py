"""Import a NordBord "Nordic" export (xlsx) into the nordico template.

Export shape (VALD NordBord): one row per (player, session) —

    Name | ExternalId | Date UTC | Time UTC | Device | Test |
    L Reps | R Reps | L Max Force (N) | R Max Force (N)

Timestamps are UTC (column names say so) and are stored as such. Same-day
retests (several rows for one player+date) collapse by PER-LEG MAX — the
"max force" semantics VALD itself reports.

Players are matched by normalized name tokens against the club's ACTIVE
players (all categories — youth train with the first team). Ambiguous or
unknown names abort the run before anything is written.

Idempotent: a (player, date) pair that already has a nordico result for
that calendar day is skipped, so re-running the same export is safe.
Results are written oldest-first so alert evaluation sees the natural
chronology.

Run (dry-run first; add --commit to write):

    docker compose exec backend python manage.py import_nordico \\
        --file /tmp/nordic.xlsx --club "Universidad de Chile" --commit
"""

from __future__ import annotations

import unicodedata
from datetime import datetime, time as time_cls, timezone as dt_timezone

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from core.models import Player
from exams.calculations import compute_result_data
from exams.models import ExamResult, ExamTemplate


def _norm(s: str) -> list[str]:
    s = unicodedata.normalize("NFD", (s or "").upper())
    return "".join(c for c in s if not unicodedata.combining(c)).split()


class Command(BaseCommand):
    help = "Import a NordBord Nordic xlsx export into the nordico template."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--club", default="Universidad de Chile")
        parser.add_argument("--commit", action="store_true",
                            help="Write results. Without it: dry-run report only.")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        template = ExamTemplate.objects.filter(
            slug="nordico", department__club__name__icontains=opts["club"],
        ).first()
        if template is None:
            raise CommandError(f"Template 'nordico' not found for club '{opts['club']}'.")

        try:
            ws = load_workbook(opts["file"], data_only=True).worksheets[0]
        except Exception as exc:  # noqa: BLE001 — surface any open/parse problem
            raise CommandError(f"Cannot read '{opts['file']}': {exc}")

        headers = [str(c).strip() if c is not None else "" for c in
                   next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col = {h: i for i, h in enumerate(headers)}
        required = ["Name", "Date UTC", "L Max Force (N)", "R Max Force (N)"]
        missing = [h for h in required if h not in col]
        if missing:
            raise CommandError(f"Missing column(s): {', '.join(missing)} — got {headers}")

        # ---- group rows by (name, date); same-day retests -> per-leg max
        groups: dict[tuple[str, object], dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[col["Name"]] or "").strip()
            if not name:
                continue
            day = row[col["Date UTC"]]
            day = day.date() if isinstance(day, datetime) else day
            left, right = row[col["L Max Force (N)"]], row[col["R Max Force (N)"]]
            if left is None or right is None:
                continue
            g = groups.setdefault(
                (name, day),
                {"left": left, "right": right,
                 "time": row[col["Time UTC"]] if "Time UTC" in col else None},
            )
            g["left"] = max(g["left"], left)
            g["right"] = max(g["right"], right)

        # ---- match players (club-wide actives) ---------------------------
        roster = list(Player.objects.filter(
            category__club__name__icontains=opts["club"], is_active=True,
        ).select_related("category"))
        resolved: dict[str, Player] = {}
        problems: list[str] = []
        for name in sorted({n for n, _ in groups}):
            tokens = _norm(name)
            cands = [p for p in roster
                     if all(t in tokens for t in _norm(f"{p.first_name} {p.last_name}"))]
            if len(cands) == 1:
                resolved[name] = cands[0]
            elif not cands:
                problems.append(f"sin match: {name!r}")
            else:
                names = ", ".join(f"{p.first_name} {p.last_name} ({p.category.name})" for p in cands)
                problems.append(f"ambiguo: {name!r} -> {names}")
        if problems:
            raise CommandError("Jugadores no resueltos:\n  " + "\n  ".join(problems))

        # ---- build + (optionally) write, oldest first --------------------
        created = skipped = 0
        with transaction.atomic():
            for (name, day), g in sorted(groups.items(), key=lambda kv: kv[0][1]):
                player = resolved[name]
                if ExamResult.objects.filter(
                    template__family_id=template.family_id, player=player,
                    recorded_at__date=day,
                ).exists():
                    skipped += 1
                    continue
                t = g["time"] if isinstance(g.get("time"), time_cls) else time_cls(12, 0)
                recorded_at = datetime.combine(day, t, tzinfo=dt_timezone.utc)
                raw = {"left_max": g["left"], "right_max": g["right"]}
                data, snapshot = compute_result_data(template, raw, player=player)
                self.stdout.write(
                    f"  {player.first_name} {player.last_name:<22} {day} | "
                    f"L {raw['left_max']} / R {raw['right_max']} | "
                    f"imb {data.get('imbalance'):+.1f}%"
                )
                if opts["commit"]:
                    # .create() (not bulk_create) so post_save signals run:
                    # player-state writeback + alert-rule evaluation.
                    ExamResult.objects.create(
                        player=player, template=template, recorded_at=recorded_at,
                        result_data=data, inputs_snapshot=snapshot,
                    )
                created += 1

        mode = "CREATED" if opts["commit"] else "would create (dry-run)"
        self.stdout.write(self.style.SUCCESS(
            f"{mode}: {created} | skipped (already imported): {skipped}"
        ))
