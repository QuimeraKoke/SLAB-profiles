"""Import the club's Maestro_CK workbook into the `ck` template.

Additive + idempotent: one ExamResult per (player, sample date); rows whose
(player, date) already exist in the platform are skipped, so re-running with
an updated Maestro file only adds the new samples. Player matching is
alias-first, then the sheet's "X.Apellido" code (initial + last name,
accent-insensitive) — same matcher as `import_lesiones`. Codes that don't
match any roster player (e.g. ex-players in the historical file) are
reported and skipped, never invented.

After commit, each affected player's LATEST CK result runs through the
band-rule evaluator (bulk_create skips signals — same trap the wellness
sync had), so reference-band alerts fire/resolve off current values only,
not historical ones.

    docker compose exec backend python manage.py import_ck \\
        --file /tmp/maestro_ck.xlsx            # dry-run: show the plan
    docker compose exec backend python manage.py import_ck \\
        --file /tmp/maestro_ck.xlsx --commit   # write
"""
from __future__ import annotations

import json
import unicodedata
from datetime import date, datetime, time
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import Category, Player, PlayerAlias
from exams.models import ExamResult, ExamTemplate


def _norm(s: str | None) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower().strip()


def _to_date(raw) -> date | None:
    if raw in (None, ""):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    try:
        return datetime.fromisoformat(str(raw).strip()[:19]).date()
    except ValueError:
        return None


class Command(BaseCommand):
    help = "Import the Maestro_CK workbook (historical CK samples) into the 'ck' template."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the .xlsx workbook.")
        parser.add_argument("--club", default="Universidad de Chile")
        parser.add_argument("--category", default="Primer Equipo")
        parser.add_argument("--sheet", default=None, help="Worksheet name (default: first).")
        parser.add_argument("--commit", action="store_true",
                            help="Write to the database (default is a dry-run plan).")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        category = Category.objects.filter(
            club__name=opts["club"], name=opts["category"],
        ).select_related("club").first()
        if category is None:
            raise CommandError(f"Category '{opts['category']}' not found in club '{opts['club']}'.")
        template = ExamTemplate.objects.filter(
            slug="ck", department__club=category.club,
        ).first()
        if template is None:
            raise CommandError("Template 'ck' not found — run seed_medico_indicators first.")

        wb = load_workbook(path, data_only=True)
        ws = wb[opts["sheet"]] if opts["sheet"] else wb[wb.sheetnames[0]]
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        col = {h: i for i, h in enumerate(headers)}
        for required in ("Jugador", "Fecha_muestra", "CK_U_L"):
            if required not in col:
                raise CommandError(f"Missing column '{required}' (found: {headers})")

        # Includes INACTIVE players on purpose: the Maestro file is
        # historical, and ex-players' samples belong on their (preserved)
        # records. Inactive players stay out of rosters/dashboards anyway.
        players = list(Player.objects.filter(category=category))
        aliases = {
            _norm(a.value): a.player_id
            for a in PlayerAlias.objects.filter(player__category=category)
        }
        by_id = {p.id: p for p in players}

        def match_player(code: str) -> Player | None:
            pid = aliases.get(_norm(code))
            if pid and pid in by_id:
                return by_id[pid]
            if "." in code:
                ini, last = code.split(".", 1)
                # Space-insensitive: "L.DiYorio" must match "DI YORIO".
                want = _norm(last).replace(" ", "")
                cands = [
                    p for p in players
                    if _norm(p.last_name).replace(" ", "").startswith(want)
                    and _norm(p.first_name).startswith(_norm(ini))
                ]
                if len(cands) == 1:
                    return cands[0]
            return None

        # Existing (player, day) keys — prefer the clinical `fecha` field,
        # falling back to recorded_at's date (team-table saves stamp
        # recorded_at at entry time, not sample time).
        existing: set[tuple] = set()
        for pid, rec, data in ExamResult.objects.filter(
            template=template, player__category=category,
        ).values_list("player_id", "recorded_at", "result_data"):
            day = (data or {}).get("fecha") or timezone.localtime(rec).date().isoformat()
            existing.add((pid, str(day)[:10]))

        plan, unmatched, skipped_dup, bad = [], {}, 0, 0
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(c not in (None, "") for c in row):
                continue
            code = str(row[col["Jugador"]] or "").strip()
            day = _to_date(row[col["Fecha_muestra"]])
            try:
                valor = float(row[col["CK_U_L"]])
            except (TypeError, ValueError):
                valor = None
            if not code or day is None or valor is None:
                bad += 1
                continue
            player = match_player(code)
            if player is None:
                unmatched[code] = unmatched.get(code, 0) + 1
                continue
            key = (player.id, day.isoformat())
            if key in existing:
                skipped_dup += 1
                continue
            existing.add(key)
            plan.append({"row": idx, "player": player, "day": day, "valor": valor})

        self.stdout.write(
            f"Plan: crear {len(plan)} resultados CK · ya existentes {skipped_dup} · "
            f"filas inválidas {bad} · códigos sin match {sum(unmatched.values())}"
        )
        for code, n in sorted(unmatched.items(), key=lambda kv: -kv[1]):
            self.stdout.write(self.style.WARNING(f"  sin match: {code} ({n} filas) — omitido"))
        per_player: dict = {}
        for p in plan:
            per_player[p["player"].last_name] = per_player.get(p["player"].last_name, 0) + 1
        self.stdout.write("  por jugador: " + ", ".join(
            f"{k} {v}" for k, v in sorted(per_player.items())))

        if not opts["commit"]:
            self.stdout.write(self.style.NOTICE("Dry-run — nada escrito. Repite con --commit."))
            return

        with transaction.atomic():
            ExamResult.objects.bulk_create([
                ExamResult(
                    player=p["player"], template=template,
                    recorded_at=timezone.make_aware(datetime.combine(p["day"], time(12, 0))),
                    result_data={"fecha": p["day"].isoformat(), "valor": p["valor"]},
                    inputs_snapshot={},
                )
                for p in plan
            ], batch_size=400)

        # Band alerts off each affected player's LATEST reading only —
        # historical rows must not fire (nor mask) current-state alerts.
        from goals.evaluator import evaluate_threshold_rules_for_result
        fired = 0
        for pid in {p["player"].id for p in plan}:
            latest = ExamResult.objects.filter(
                player_id=pid, template=template,
            ).order_by("-recorded_at").first()
            if latest is not None:
                fired += len(evaluate_threshold_rules_for_result(latest))

        log_dir = Path(__file__).resolve().parents[3] / "migration_runs"
        log_dir.mkdir(exist_ok=True)
        log_path = log_dir / f"ck-import-{timezone.now().strftime('%Y%m%dT%H%M%S')}.jsonl"
        with log_path.open("w") as fh:
            fh.write(json.dumps({"kind": "header", "file": path.name, "created": len(plan),
                                 "skipped_existing": skipped_dup, "unmatched": unmatched}) + "\n")
            for p in plan:
                fh.write(json.dumps({
                    "row": p["row"], "player": f"{p['player'].first_name} {p['player'].last_name}",
                    "fecha": p["day"].isoformat(), "valor": p["valor"],
                }) + "\n")

        self.stdout.write(self.style.SUCCESS(
            f"Listo: {len(plan)} resultados CK creados · {fired} alertas de banda "
            f"evaluadas sobre las últimas tomas. Log: {log_path.name}"
        ))
