"""Import the club's injury-surveillance sheet (Vigilancia de Lesiones).

Reads the 'Registro' sheet of the Fuller-format surveillance workbook and
creates one Episode per row on the 'lesiones' episodic template:

- **Alta** rows get TWO results — the opening one at the injury date
  (stage=injured) and the closing one at the return date (stage=closed,
  `actual_return_date` set) — so `Episode.ended_at` and the closed status
  derive correctly from the lifecycle signal.
- **Activa** rows get ONE result (stage=injured) → the player's status
  flips to Lesionado and he surfaces in the Daily.

Player matching: `PlayerAlias` first, then the sheet's "X.Apellido" code
(first-name initial + last name, accent-insensitive).

`--wipe` deletes ALL existing lesiones episodes + results for the category
first (the sheet is the club's single source of truth for injuries).
Dry-run by default; pass `--commit` to write. Every run appends a JSONL
log under `backend/migration_runs/`.

    docker compose exec backend python manage.py import_lesiones \\
        --file /tmp/lesiones.xlsx --wipe            # dry-run: show the plan
    docker compose exec backend python manage.py import_lesiones \\
        --file /tmp/lesiones.xlsx --wipe --commit   # replace the data
"""
from __future__ import annotations

import json
import unicodedata
from datetime import date, datetime, time
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import Category, Player, PlayerAlias
from exams.models import Episode, ExamResult, ExamTemplate

SHEET = "Registro"

# Sheet header → position lookup is done by name, so column reordering in
# future workbook versions doesn't silently corrupt the mapping.
REQUIRED_HEADERS = ["Jugador", "Fecha lesion (inicio)", "Region", "Estado"]

# "Tratamiento: Kinesico" prefixes inside Notas → the template's categorical.
_TRATAMIENTO_MAP = {
    "kinesico": "Kinésico",
    "kinesico + quirurgico": "Kinésico + quirúrgico",
    "quirurgico": "Kinésico + quirúrgico",
    "reposo": "Reposo deportivo",
    "reposo deportivo": "Reposo deportivo",
}


def _norm(s: str | None) -> str:
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower().strip()


def _to_date(raw) -> date | None:
    if raw in (None, ""):
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    try:
        return datetime.fromisoformat(str(raw).strip()).date()
    except ValueError:
        return None


def _to_int(raw) -> int | None:
    try:
        return int(float(raw))
    except (TypeError, ValueError):
        return None


def _aware(d: date) -> datetime:
    return timezone.make_aware(datetime.combine(d, time(12, 0)))


class Command(BaseCommand):
    help = "Import the Vigilancia de Lesiones workbook into the 'lesiones' template."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True, help="Path to the .xlsx workbook.")
        parser.add_argument("--club", default="Universidad de Chile")
        parser.add_argument("--category", default="Primer Equipo")
        parser.add_argument("--sheet", default=SHEET)
        parser.add_argument("--wipe", action="store_true",
                            help="Delete every existing lesiones episode/result for the category first.")
        parser.add_argument("--commit", action="store_true",
                            help="Write to the database (default is a dry-run plan).")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise CommandError("openpyxl is required") from exc

        path = Path(opts["file"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        category = Category.objects.filter(
            club__name=opts["club"], name=opts["category"],
        ).select_related("club").first()
        if category is None:
            raise CommandError(f"Category '{opts['category']}' not found in club '{opts['club']}'.")

        template = ExamTemplate.objects.filter(
            slug="lesiones", department__club=category.club,
        ).first()
        if template is None:
            raise CommandError("Template 'lesiones' not found — run seed_lesiones first.")

        wb = load_workbook(path, data_only=True)
        if opts["sheet"] not in wb.sheetnames:
            raise CommandError(f"Sheet '{opts['sheet']}' not in workbook ({wb.sheetnames}).")
        ws = wb[opts["sheet"]]
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        missing = [h for h in REQUIRED_HEADERS if h not in headers]
        if missing:
            raise CommandError(f"Missing expected columns: {missing}")
        col = {h: i for i, h in enumerate(headers)}

        players = list(Player.objects.filter(category=category, is_active=True))
        aliases = {
            _norm(a.value): a.player_id
            for a in PlayerAlias.objects.filter(player__category=category)
        }
        by_id = {p.id: p for p in players}

        def match_player(code: str) -> Player | None:
            pid = aliases.get(_norm(code))
            if pid and pid in by_id:
                return by_id[pid]
            if "." in code:
                ini, last = code.split(".", 1)
                cands = [
                    p for p in players
                    if _norm(p.last_name).startswith(_norm(last))
                    and _norm(p.first_name).startswith(_norm(ini))
                ]
                if len(cands) == 1:
                    return cands[0]
            return None

        # ── Parse rows into a plan ────────────────────────────────────
        plan, unmatched, bad_rows = [], [], []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(c not in (None, "") for c in row):
                continue

            def v(header):
                raw = row[col[header]] if header in col else None
                return raw if raw not in (None, "") else None

            code = str(v("Jugador") or "").strip()
            diagnosed = _to_date(v("Fecha lesion (inicio)"))
            if not code or diagnosed is None:
                bad_rows.append({"row": idx, "reason": "sin jugador o fecha de lesión"})
                continue
            player = match_player(code)
            if player is None:
                unmatched.append({"row": idx, "code": code})
                continue

            estado = str(v("Estado") or "").strip().lower()
            returned = _to_date(v("Fecha alta (retorno)"))
            today = timezone.localdate()
            # "Alta" with a FUTURE return date = scheduled discharge — the
            # player is still out. Keep the episode OPEN with that date as
            # the expected return (the Daily shows "Retorno estimado"); it
            # closes for real once the date has passed in a later import.
            scheduled = returned if (returned and returned > today) else None
            closed = estado != "activa" and scheduled is None
            notes = str(v("Notas / Observaciones") or "").strip()

            data = {
                "diagnosed_at": diagnosed.isoformat(),
                "type": v("Tipo / Diagnostico (Fuller)"),
                "body_part": v("Region"),
                "lado": v("Lado"),
                "body_part_detail": str(v("Localizacion especifica") or "").strip() or None,
                "severity": v("Severidad (Fuller)"),
                "bamic": str(v("BAMIC (RM)")).strip() if v("BAMIC (RM)") is not None else None,
                "hallazgos_rm": str(v("Hallazgos RM / Informe") or "").strip() or None,
                "exposicion": v("Contexto"),
                "mecanismo": v("Mecanismo"),
                "modo": v("Modo"),
                "recurrencia": v("Recurrencia"),
                "tipo_recurrencia": v("Tipo recurrencia"),
                "dias_perdidos": _to_int(v("Dias de baja")),
                "partidos_perdidos": _to_int(v("Partidos perdidos")),
                "notes": notes or None,
            }
            # "Tratamiento: Kinesico" convention inside Notas → categorical.
            if _norm(notes).startswith("tratamiento:"):
                mapped = _TRATAMIENTO_MAP.get(_norm(notes.split(":", 1)[1]))
                if mapped:
                    data["tratamiento"] = mapped
            if scheduled:
                data["expected_return_date"] = scheduled.isoformat()
            data = {k: val for k, val in data.items() if val is not None}

            plan.append({
                "row": idx,
                "external_id": _to_int(v("ID")),
                "code": code,
                "player": player,
                "diagnosed": diagnosed,
                "returned": returned if closed else None,
                "closed": closed,
                "data": data,
                "raw": {h: (str(row[i]) if row[i] is not None else None) for h, i in col.items()},
            })

        # ── Report the plan ───────────────────────────────────────────
        n_open = sum(1 for p in plan if not p["closed"])
        existing_results = ExamResult.objects.filter(
            template=template, player__category=category,
        ).count()
        existing_episodes = Episode.objects.filter(
            template=template, player__category=category,
        ).count()

        self.stdout.write(
            f"Plan: {len(plan)} lesiones ({n_open} activas, {len(plan) - n_open} cerradas) "
            f"para {category.club.name} / {category.name}."
        )
        if opts["wipe"]:
            self.stdout.write(self.style.WARNING(
                f"--wipe: se eliminarán {existing_episodes} episodios y "
                f"{existing_results} resultados 'lesiones' existentes."
            ))
        for u in unmatched:
            self.stdout.write(self.style.ERROR(
                f"  fila {u['row']}: jugador '{u['code']}' sin match — fila omitida"
            ))
        for b in bad_rows:
            self.stdout.write(self.style.ERROR(f"  fila {b['row']}: {b['reason']} — fila omitida"))
        for p in plan:
            if p["closed"]:
                dest = str(p["returned"] or "—")
            elif p["data"].get("expected_return_date"):
                dest = f"ABIERTA · alta prog. {p['data']['expected_return_date']}"
            else:
                dest = "ACTIVA"
            self.stdout.write(
                f"  fila {p['row']:>3}: {p['player'].first_name} {p['player'].last_name:<14} "
                f"{p['data'].get('body_part_detail') or p['data'].get('type', '?'):<38} "
                f"{p['diagnosed']} → {dest}"
            )

        if not opts["commit"]:
            self.stdout.write(self.style.NOTICE("Dry-run — nada escrito. Repite con --commit."))
            return
        if unmatched:
            raise CommandError(
                "Hay jugadores sin match — agrega un PlayerAlias y reintenta. Nada se escribió."
            )

        # ── Write ─────────────────────────────────────────────────────
        with transaction.atomic():
            if opts["wipe"]:
                ExamResult.objects.filter(template=template, player__category=category).delete()
                Episode.objects.filter(template=template, player__category=category).delete()

            created = 0
            for p in plan:
                episode = Episode.objects.create(
                    player=p["player"],
                    template=template,
                    started_at=_aware(p["diagnosed"]),
                    legacy_raw={"source": path.name, "row": p["row"],
                                "external_id": p["external_id"], "data": p["raw"]},
                )
                opening = dict(p["data"], stage="injured")
                ExamResult.objects.create(
                    player=p["player"], template=template, episode=episode,
                    recorded_at=_aware(p["diagnosed"]), result_data=opening,
                    inputs_snapshot={},
                )
                if p["closed"]:
                    closing = dict(p["data"], stage="closed")
                    if p["returned"]:
                        closing["actual_return_date"] = p["returned"].isoformat()
                    ExamResult.objects.create(
                        player=p["player"], template=template, episode=episode,
                        recorded_at=_aware(p["returned"] or p["diagnosed"]),
                        result_data=closing, inputs_snapshot={},
                    )
                created += 1

            # Statuses: result saves already recomputed the players we touched,
            # but wiped players with no new injury need an explicit pass.
            from exams.episode_lifecycle import recompute_player_status
            for player in players:
                recompute_player_status(player)

        # ── Run log (repo convention: migration_runs/*.jsonl) ─────────
        log_dir = Path(__file__).resolve().parents[3] / "migration_runs"
        log_dir.mkdir(exist_ok=True)
        stamp = timezone.now().strftime("%Y%m%dT%H%M%S")
        log_path = log_dir / f"lesiones-import-{stamp}.jsonl"
        with log_path.open("w") as fh:
            fh.write(json.dumps({
                "kind": "header", "file": path.name, "category": category.name,
                "wiped": opts["wipe"], "episodes": created,
                "wiped_episodes": existing_episodes if opts["wipe"] else 0,
            }) + "\n")
            for p in plan:
                fh.write(json.dumps({
                    "row": p["row"], "external_id": p["external_id"],
                    "player": f"{p['player'].first_name} {p['player'].last_name}",
                    "closed": p["closed"], "diagnosed": p["diagnosed"].isoformat(),
                    "returned": p["returned"].isoformat() if p["returned"] else None,
                }) + "\n")

        self.stdout.write(self.style.SUCCESS(
            f"Listo: {created} episodios creados ({n_open} activos). Log: {log_path.name}"
        ))
