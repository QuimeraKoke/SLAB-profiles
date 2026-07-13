"""Import a force-plate "IMTP" export (xlsx) into the imtp template.

Export shape (VALD ForceDecks style): one row per (player, session, trial) —

    Name | ExternalId | Test Type | Date | Time | BW [KG] | Reps | Tags |
    Peak Vertical Force [N] | Peak Vertical Force / BM [N/kg] | RFD - 200ms [N/s]

Header cells may carry trailing spaces (they do in real exports) — headers
are stripped before matching. Multi-trial days collapse to the trial with
the highest Peak Vertical Force, keeping THAT trial's BM-relative force and
RFD (whole-row, so the numbers stay internally coherent).

Players are matched by normalized name tokens against the club's ACTIVE
players (all categories — youth train with the first team). Ambiguous or
unknown names abort the run before anything is written.

Idempotent: a (player, date) pair that already has an imtp result for that
calendar day is skipped, so re-running the same export is safe.

Run (dry-run first; add --commit to write):

    docker compose exec backend python manage.py import_imtp \\
        --file /tmp/imtp.xlsx --club "Universidad de Chile" --commit
"""

from __future__ import annotations

import unicodedata
from datetime import datetime, time as time_cls

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone

from core.models import Player
from exams.calculations import compute_result_data
from exams.models import ExamResult, ExamTemplate


def _norm(s: str) -> list[str]:
    s = unicodedata.normalize("NFD", (s or "").upper())
    return "".join(c for c in s if not unicodedata.combining(c)).split()


class Command(BaseCommand):
    help = "Import a force-plate IMTP xlsx export into the imtp template."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--club", default="Universidad de Chile")
        parser.add_argument("--commit", action="store_true",
                            help="Write results. Without it: dry-run report only.")

    def handle(self, *args, **opts):
        from openpyxl import load_workbook

        template = ExamTemplate.objects.filter(
            slug="imtp", department__club__name__icontains=opts["club"],
        ).first()
        if template is None:
            raise CommandError(f"Template 'imtp' not found for club '{opts['club']}'.")

        try:
            ws = load_workbook(opts["file"], data_only=True).worksheets[0]
        except Exception as exc:  # noqa: BLE001 — surface any open/parse problem
            raise CommandError(f"Cannot read '{opts['file']}': {exc}")

        headers = [str(c).strip() if c is not None else "" for c in
                   next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        col = {h: i for i, h in enumerate(headers)}
        required = ["Name", "Date", "Peak Vertical Force [N]",
                    "Peak Vertical Force / BM [N/kg]", "RFD - 200ms [N/s]"]
        missing = [h for h in required if h not in col]
        if missing:
            raise CommandError(f"Missing column(s): {', '.join(missing)} — got {headers}")

        # ---- group by (name, date); multi-trial -> best whole trial ------
        groups: dict[tuple[str, object], dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[col["Name"]] or "").strip()
            if not name:
                continue
            day = row[col["Date"]]
            day = day.date() if isinstance(day, datetime) else day
            peak = row[col["Peak Vertical Force [N]"]]
            if peak is None:
                continue
            trial = {
                "peak_vertical_force": peak,
                "peak_force_bodymass": row[col["Peak Vertical Force / BM [N/kg]"]],
                "rfd_200ms": row[col["RFD - 200ms [N/s]"]],
                "time": row[col["Time"]] if "Time" in col else None,
            }
            key = (name, day)
            if key not in groups or trial["peak_vertical_force"] > groups[key]["peak_vertical_force"]:
                groups[key] = trial

        # ---- match players (club-wide actives) ---------------------------
        roster = list(Player.objects.filter(
            category__club__name__icontains=opts["club"], is_active=True,
        ).select_related("category"))
        resolved: dict[str, Player] = {}
        problems: list[str] = []
        for name in sorted({n for n, _ in groups}):
            tokens = _norm(name)
            cands = [p for p in roster
                     if all(t in tokens for t in _norm(f"{p.first_name} {p.last_name}"))]
            if len(cands) == 1:
                resolved[name] = cands[0]
            elif not cands:
                problems.append(f"sin match: {name!r}")
            else:
                names = ", ".join(f"{p.first_name} {p.last_name} ({p.category.name})" for p in cands)
                problems.append(f"ambiguo: {name!r} -> {names}")
        if problems:
            raise CommandError("Jugadores no resueltos:\n  " + "\n  ".join(problems))

        # ---- build + (optionally) write, oldest first --------------------
        tz = timezone.get_default_timezone()
        created = skipped = 0
        with transaction.atomic():
            for (name, day), g in sorted(groups.items(), key=lambda kv: kv[0][1]):
                player = resolved[name]
                if ExamResult.objects.filter(
                    template__family_id=template.family_id, player=player,
                    recorded_at__date=day,
                ).exists():
                    skipped += 1
                    continue
                t = g["time"] if isinstance(g.get("time"), time_cls) else time_cls(12, 0)
                recorded_at = timezone.make_aware(datetime.combine(day, t), tz)
                raw = {k: g[k] for k in
                       ("peak_vertical_force", "peak_force_bodymass", "rfd_200ms")}
                data, snapshot = compute_result_data(template, raw, player=player)
                self.stdout.write(
                    f"  {player.first_name} {player.last_name:<22} {day} | "
                    f"peak {raw['peak_vertical_force']} N | "
                    f"{raw['peak_force_bodymass']} N/kg | RFD {raw['rfd_200ms']} N/s"
                )
                if opts["commit"]:
                    # .create() (not bulk_create) so post_save signals run.
                    ExamResult.objects.create(
                        player=player, template=template, recorded_at=recorded_at,
                        result_data=data, inputs_snapshot=snapshot,
                    )
                created += 1

        mode = "CREATED" if opts["commit"] else "would create (dry-run)"
        self.stdout.write(self.style.SUCCESS(
            f"{mode}: {created} | skipped (already imported): {skipped}"
        ))
