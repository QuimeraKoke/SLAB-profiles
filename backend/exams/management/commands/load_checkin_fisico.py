"""Load wellness Check-IN responses from a Google-Forms .xlsx into the
`checkin_fisico` template, via the shared `wellness_ingest` pipeline.

    docker compose exec backend python manage.py load_checkin_fisico \\
        --file /tmp/checkin.xlsx --club "Universidad de Chile" --category "Primer Equipo"

Idempotent on (player, recorded_at); molestia codes are normalized + validated
against the template glossary, and molestia / estado-mismatch alerts are raised.
For the automated Google-Sheet sync, see `exams.tasks.sync_wellness_responses`.
"""
from __future__ import annotations

from django.core.management.base import BaseCommand, CommandError

_SHEET = "Respuestas de formulario 1"


class Command(BaseCommand):
    help = "Load wellness Check-IN responses from an .xlsx into checkin_fisico."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--club", default="Universidad de Chile")
        parser.add_argument("--category", default="Primer Equipo")
        parser.add_argument("--sheet", default=_SHEET)
        parser.add_argument("--mode", default="all", choices=["all", "today", "reconcile"])
        parser.add_argument("--since-days", type=int, default=3)
        parser.add_argument("--dry-run", action="store_true")

    def handle(self, *args, **opts):
        import openpyxl

        from core.models import Category
        from exams.models import ExamTemplate
        from exams.wellness_ingest import WELLNESS_SLUG, ingest_wellness

        cat = (
            Category.objects.filter(name=opts["category"], club__name=opts["club"])
            .select_related("club").first()
        )
        if cat is None:
            raise CommandError(f"Category '{opts['category']}' not found in club '{opts['club']}'.")
        template = ExamTemplate.objects.filter(
            slug=WELLNESS_SLUG, department__club=cat.club,
        ).first()
        if template is None:
            raise CommandError("Template 'checkin_fisico' not found — run seed_checkin_fisico first.")

        wb = openpyxl.load_workbook(opts["file"], data_only=True, read_only=True)
        try:
            ws = wb[opts["sheet"]]
        except KeyError:
            raise CommandError(f"Sheet '{opts['sheet']}' not found. Available: {wb.sheetnames}")
        it = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else "" for c in next(it)]
        except StopIteration:
            raise CommandError("Empty sheet.")
        rows = []
        for raw in it:
            if not any(c is not None and str(c).strip() for c in raw):
                continue
            rows.append({h: v for h, v in zip(header, raw) if h})

        report = ingest_wellness(
            rows, template=template, category=cat,
            mode=opts["mode"], since_days=opts["since_days"], dry_run=opts["dry_run"],
        )
        tag = "DRY-RUN " if opts["dry_run"] else ""
        self.stdout.write(self.style.SUCCESS(
            f"{tag}Check-IN: +{report['created']} created, {report['skipped']} skipped "
            f"(of {report['rows']} rows) · molestias={report['molestias']} "
            f"mismatches={report['mismatches']} alerts={report['alerts']}"
        ))
        if report["unmatched"]:
            self.stdout.write(self.style.WARNING("Unmatched players (not in active roster):"))
            for n, c in sorted(report["unmatched"].items(), key=lambda x: -x[1]):
                self.stdout.write(f"  {n} ({c} rows)")
        if report["unknown_codes"]:
            self.stdout.write(self.style.WARNING(
                "Unknown molestia codes (not in glossary): "
                + ", ".join(f"{k}×{v}" for k, v in sorted(report["unknown_codes"].items()))
            ))
