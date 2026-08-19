"""Parse + ingest the Pentacompartimental (5-component anthropometry) workbook.

Engine behind the `import_pentacompartimental` management command — a CLI-only
path for bulk-loading an ISAK export (historical backfills, or a season's worth
of assessments arriving in one file). Routine capture goes through the
template's own `team_table` / `bulk_ingest` modes on /subir-datos instead; a
self-service upload for this format existed briefly and was removed as
redundant.

The workbook is the ISAK software's "Modelo 5 componentes" export: four title /
group-header rows, the real header on row 5, data from row 6. Each player is a
BLOCK — the name appears once in column A and is blank on the following rows —
and every row in the block is one assessment, dated in column B ("Informes").
Only the 27 raw measurements (columns 4-30) are read: the export's ~50 computed
columns are recalculated here from the template's own formulas, and `sexo` comes
from the player record, not the sheet.

Additive by design: dedup is on (player, assessment date) and an existing result
is NEVER overwritten, so re-uploading the same report is a no-op.

`bulk_create` emits no post_save signals, so the band-alert evaluation a
registrar save gets for free is invoked explicitly — but only for assessments
inside the `ALERT_STALE_DAYS` window, since an alert anchored on an older
reading would just be expired again by the staleness sweep. That keeps a
historical backfill from flooding the alert list.
"""
from __future__ import annotations

import io
import unicodedata
from datetime import datetime, timedelta

from django.db import transaction
from django.utils import timezone

from core.models import Player
from exams.calculations import compute_result_data
from exams.models import ExamResult

SHEET = "Modelo 5 componentes"
FIRST_DATA_ROW = 6
NAME_COL = 1          # "PACIENTES" — forward-filled per block
DATE_COL = 2          # "Informes"  — the assessment date

# Excel column (1-based) → template raw field key. The template's other fields
# are either calculated (recomputed) or not present in the export.
RAW_MAP = {
    4: "peso", 5: "talla", 6: "talla_sentado",
    7: "biacromial", 8: "diam_torax_transverso", 9: "diam_torax_ap",
    10: "bi_iliocrestideo", 11: "humero", 12: "femur",
    13: "perim_cabeza", 14: "perim_brazo_relajado", 15: "perim_brazo_contraido",
    16: "perim_antebrazo", 17: "perim_torax", 18: "cintura", 19: "caderas",
    20: "muslo_gluteo", 21: "muslo_medio", 22: "pierna_perim",
    23: "pliegue_bicipital", 24: "pliegue_triceps", 25: "pliegue_subescapular",
    26: "pliegue_supracrestideo", 27: "pliegue_supra", 28: "pliegue_abdomen",
    29: "pliegue_muslo", 30: "pliegue_pierna",
}

MIN_NAME_TOKEN_HITS = 2   # below this the name is treated as unmatched


class PentaParseError(ValueError):
    """The upload isn't a readable 5-component workbook."""


# ---------- name matching ----------

def _tokens(s: str) -> list[str]:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode().lower()
    return [t for t in s.replace(".", " ").replace(",", " ").split() if len(t) > 1]


def _tok_match(a: str, b: str) -> bool:
    """Equal, or one is a ≥4-char prefix of the other (Bolaño/Bolaños)."""
    if a == b:
        return True
    if len(a) >= 4 and b.startswith(a):
        return True
    if len(b) >= 4 and a.startswith(b):
        return True
    return False


def build_matcher(club):
    """Name → Player resolver over the club's ENTIRE roster.

    Deliberately not category-scoped: youth players are listed alongside the
    first team in these reports. Ties break toward Primer Equipo.
    """
    cand = [
        (p, _tokens(f"{p.first_name} {p.last_name}"))
        for p in Player.objects.filter(category__club=club).select_related("category")
    ]

    def match(name: str):
        et = _tokens(name)
        best, best_score = None, (-1, -1)
        for p, pt in cand:
            score = sum(1 for a in et if any(_tok_match(a, b) for b in pt))
            pe = 1 if p.category and p.category.name == "Primer Equipo" else 0
            if (score, pe) > best_score:
                best_score, best = (score, pe), p
        return best if best_score[0] >= MIN_NAME_TOKEN_HITS else None

    return match


# ---------- cell coercion ----------

def _num(v):
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def _to_date(v):
    if v in (None, ""):
        return None
    if hasattr(v, "date"):
        return v.date()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


# ---------- parse ----------

def parse_workbook(file_bytes: bytes, sheet: str = SHEET) -> list[tuple]:
    """Data rows (row 6 onward) of the 5-component sheet. Raises PentaParseError."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - dependency is installed
        raise PentaParseError(f"openpyxl no disponible: {exc}") from exc
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise PentaParseError(
            f"No se pudo leer el archivo (¿es un .xlsx válido?): {exc}"
        ) from exc
    if sheet not in wb.sheetnames:
        raise PentaParseError(
            f"El archivo no tiene la hoja '{sheet}'. Hojas encontradas: "
            f"{', '.join(wb.sheetnames) or '(ninguna)'}."
        )
    return list(wb[sheet].iter_rows(min_row=FIRST_DATA_ROW, values_only=True))


# ---------- ingest ----------

def run(
    file_bytes: bytes,
    *,
    club,
    template,
    dry_run: bool = True,
    sheet: str = SHEET,
    fire_alerts: bool = True,
) -> dict:
    """Plan (and optionally write) results from the workbook. JSON-friendly report.

    Additive: an existing (player, date) is reported as skipped, never updated.
    """
    rows = parse_workbook(file_bytes, sheet=sheet)
    match = build_matcher(club)

    existing = {
        (pid, dt.date())
        for pid, dt in ExamResult.objects.filter(template=template)
        .values_list("player_id", "recorded_at")
    }

    per_player: dict = {}
    unmatched: dict[str, int] = {}
    to_create: list[ExamResult] = []
    seen_in_file: set[tuple] = set()
    cur_name = None
    skipped_existing = skipped_dupe = skipped_incomplete = skipped_undated = 0

    for r in rows:
        # Column A carries the name only on a block's first row.
        if len(r) > 0 and r[NAME_COL - 1] not in (None, ""):
            cur_name = str(r[NAME_COL - 1]).strip()
        if cur_name is None:
            continue
        d = _to_date(r[DATE_COL - 1] if len(r) >= DATE_COL else None)
        if d is None:
            skipped_undated += 1
            continue
        player = match(cur_name)
        if player is None:
            unmatched[cur_name] = unmatched.get(cur_name, 0) + 1
            continue
        key = (player.id, d)
        if key in existing:
            skipped_existing += 1
            continue
        if key in seen_in_file:
            skipped_dupe += 1
            continue

        raw = {
            k: _num(r[c - 1])
            for c, k in RAW_MAP.items()
            if len(r) >= c and _num(r[c - 1]) is not None
        }
        # Peso + talla drive every mass equation; without them the row is a
        # spacer or a partially-filled draft, not an assessment.
        if not raw.get("peso") or not raw.get("talla"):
            skipped_incomplete += 1
            continue
        seen_in_file.add(key)
        raw["sexo"] = 1 if (getattr(player, "sex", "M") or "M").upper().startswith("M") else 2
        result_data, inputs_snapshot = compute_result_data(template, raw, player=player)
        to_create.append(ExamResult(
            player=player, template=template,
            recorded_at=timezone.make_aware(
                datetime.combine(d, datetime.min.time()),
                timezone.get_default_timezone(),
            ),
            result_data=result_data, inputs_snapshot=inputs_snapshot,
        ))
        slot = per_player.setdefault(player.id, {
            "excel_name": cur_name,
            "player": f"{player.first_name} {player.last_name}",
            "player_id": str(player.id),
            "category": player.category.name if player.category else None,
            "new": 0, "dates": [],
        })
        slot["new"] += 1
        slot["dates"].append(d.isoformat())

    alerts_fired = 0
    if not dry_run and to_create:
        with transaction.atomic():
            ExamResult.objects.bulk_create(to_create, batch_size=500)
            if fire_alerts:
                alerts_fired = _fire_band_alerts(to_create)

    return {
        "club": club.name,
        "template": template.name,
        "sheet": sheet,
        "rows_read": len(rows),
        "created": len(to_create),
        "skipped_existing": skipped_existing,
        "skipped_duplicate_in_file": skipped_dupe,
        "skipped_incomplete": skipped_incomplete,
        "skipped_undated": skipped_undated,
        "players": sorted(per_player.values(), key=lambda s: s["player"]),
        "unmatched": [
            {"name": n, "rows": c} for n, c in sorted(unmatched.items())
        ],
        "alerts_fired": alerts_fired,
        "committed": not dry_run,
        "dry_run": dry_run,
    }


def _fire_band_alerts(results: list[ExamResult]) -> int:
    """Run the template's band rules over freshly created results.

    Only assessments within ALERT_STALE_DAYS are evaluated: an alert anchored on
    an older reading is expired again by `expire_stale_alerts` on the next
    sweep, so firing it would just churn the alert list during a backfill.
    """
    from goals.evaluator import ALERT_STALE_DAYS, evaluate_threshold_rules_for_result

    cutoff = timezone.now() - timedelta(days=ALERT_STALE_DAYS)
    fired = 0
    for result in results:
        if result.recorded_at < cutoff:
            continue
        fired += len(evaluate_threshold_rules_for_result(result))
    return fired
